# -*- coding: utf-8 -*-
"""
scan_predictions.py
-------------------
Reads all player prediction xlsx-files from xlsx/ and writes
gData/predictions.json.

Group stage scores are read from cols F/G (0-indexed 5/6), rows 7-78.

Knockout predictions are read directly from the dedicated bracket columns,
where each match occupies two consecutive rows (home team, away team):

  BL (col 64) — Round of 32   : rows 10-11, 14-15, ..., 70-71  (16 matches)
  BS (col 71) — Round of 16   : rows 12-13, 20-21, ..., 44-45  ( 8 matches)  [5 shown here, read all non-null]
  BZ (col 78) — Quarter-finals: rows 16-17, 32-33, 48-49       ( up to 4 matches)
  CG (col 85) — Semi-finals   : rows 23-24                      ( 1 match shown; read all)
  CN (col 92) — Final & Bronze: rows 37-38 (final), 48-49 (bronze)

Usage
-----
python py/scan_predictions.py                   # project defaults
python py/scan_predictions.py xlsx/ out.json    # explicit paths
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
import pandas as pd
from pyprojroot.criterion import has_dir
from pyprojroot.root import find_root


# ---------------------------------------------------------------------------
# Project root
# ---------------------------------------------------------------------------
def project_root() -> Path:
    return find_root(has_dir(".git"), start=Path(__file__).resolve().parent)


# ---------------------------------------------------------------------------
# Column indices (1-based, as openpyxl uses)
# ---------------------------------------------------------------------------
# Each round: team col, then FT score, ET score, PEN score in the next 3 cols.
# first_rows = the home-team row of each match (away team is always row+1).
ROUNDS = {
    "r32":    ("BL", "BM", "BN", "BO", list(range(10, 72, 4))),   # 16 matches
    "r16":    ("BS", "BT", "BU", "BV", list(range(12, 72, 8))),   # 8 matches
    "qf":     ("BZ", "CA", "CB", "CC", list(range(16, 80, 16))),  # 4 matches (was range(16,56,16) — missed row 64)
    "sf":     ("CG", "CH", "CI", "CJ", [23, 55]),                 # 2 matches (was [23,39] — row 39 is wrong)
    "final":  ("CN", "CO", "CP", "CQ", [37]),
    "bronze": ("CN", "CO", "CP", "CQ", [48]),
}
ROUNDS_CI = {
    k: tuple(column_index_from_string(c) if isinstance(c, str) else c for c in v)
    for k, v in ROUNDS.items()
}

# Map (round_key, first_row) -> fasit match_id, so score.py can join by match number.
# Match ids come from col BK/BR/BY/CF/CM (one column left of the team column),
# sitting one row above each home-team row.
MATCH_ID_MAP: dict[tuple[str, int], int] = {
    # Round of 32 (matches 73-88)
    ("r32", 10): 74, ("r32", 14): 77, ("r32", 18): 73, ("r32", 22): 75,
    ("r32", 26): 83, ("r32", 30): 84, ("r32", 34): 81, ("r32", 38): 82,
    ("r32", 42): 76, ("r32", 46): 78, ("r32", 50): 79, ("r32", 54): 80,
    ("r32", 58): 86, ("r32", 62): 88, ("r32", 66): 85, ("r32", 70): 87,
    # Round of 16 (matches 89-96)
    ("r16", 12): 89, ("r16", 20): 90, ("r16", 28): 93, ("r16", 36): 94,
    ("r16", 44): 91, ("r16", 52): 92, ("r16", 60): 95, ("r16", 68): 96,
    # Quarterfinals (matches 97-100)
    ("qf",  16): 97, ("qf",  32): 98, ("qf",  48): 99, ("qf",  64): 100,
    # Semi-Finals (matches 101-102)
    ("sf",  23): 101, ("sf",  55): 102,
    # Final & Bronze
    ("final",  37): 104,
    ("bronze", 48): 103,
}

# Map round key -> stage label expected by score.py's score_knockout()
STAGE_LABEL = {
    "r32":    "Round of 32",
    "r16":    "Round of 16",
    "qf":     "Quarterfinals",
    "sf":     "Semi-Finals",
    "final":  "Final",
    "bronze": "Third-Place Play-Off",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _val(ws, row, col):
    """Return cell value, or None if empty/zero."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, (int, float)) and v == 0:
        return None
    if isinstance(v, str) and v.strip() in ('', '0'):
        return None
    return str(v).strip() if isinstance(v, str) else v


def _int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _score(ws, row, col):
    """Read a goal count — allows 0, returns None only for truly empty cells."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _read_matches(ws, round_key, col_team, col_ft, col_et, col_pen, first_rows):
    """Read knockout matches: teams + FT/ET/PEN scores, tagged with match_id and stage."""
    matches = []
    for r in first_rows:
        home = _val(ws, r,     col_team)
        away = _val(ws, r + 1, col_team)
        if not home and not away:
            continue
        ft_h = _score(ws, r,     col_ft)
        ft_a = _score(ws, r + 1, col_ft)
        et_h = _score(ws, r,     col_et)
        et_a = _score(ws, r + 1, col_et)
        pn_h = _score(ws, r,     col_pen)
        pn_a = _score(ws, r + 1, col_pen)
        matches.append({
            "match":  MATCH_ID_MAP.get((round_key, r)),
            "stage":  STAGE_LABEL[round_key],
            "team1":  home,
            "team2":  away,
            "ft":     [ft_h, ft_a],
            "et":     [et_h, et_a] if et_h is not None else None,
            "pen":    [pn_h, pn_a] if pn_h is not None else None,
        })
    return matches


# ---------------------------------------------------------------------------
# Read a single player xlsx
# ---------------------------------------------------------------------------
def read_player_file(path: Path, match_index: dict) -> dict:
    # Use openpyxl with data_only=True to get formula-evaluated values
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['2026 World Cup']
    warnings = []

    # --- Group stage: rows 7-78, cols F(6)/G(7) for scores ---
    group_stage = []
    for row in range(7, 79):
        raw_id = _val(ws, row, 1)   # col A = match id
        if raw_id is None:
            continue
        match_id = _int(raw_id)
        if match_id is None:
            continue

        home_goals = _score(ws, row, 6)   # col F
        away_goals = _score(ws, row, 7)   # col G

        if home_goals is None or away_goals is None:
            warnings.append(f"M{match_id}: missing score (row {row})")

        fixture = match_index.get(match_id, {})
        group_stage.append({
            "match_id":   match_id,
            "home":       fixture.get("home"),
            "away":       fixture.get("away"),
            "home_goals": home_goals,
            "away_goals": away_goals,
        })

    # --- Knockout ---
    def _rm(key):
        ct, ft, et, pn, rows = ROUNDS_CI[key]
        return _read_matches(ws, key, ct, ft, et, pn, rows)

    # Flat list of all KO matches in chronological (match number) order,
    # matching the format score.py's score_knockout() expects:
    #   [{match, stage, team1, team2, ft, et, pen}, ...]
    ko_rounds = ["r32", "r16", "qf", "sf", "final", "bronze"]
    knockout_flat = []
    knockout_by_round = {}
    for key in ko_rounds:
        matches = _rm(key)
        knockout_by_round[key] = matches
        knockout_flat.extend(matches)
    knockout_flat.sort(key=lambda m: m["match"] if m["match"] else 999)

    # Derive champion / runner-up / third-place from predicted bracket results
    def _winner(match):
        """Return the predicted winner of a match, or None if not filled in."""
        ft  = match.get("ft",  [None, None]) or [None, None]
        et  = match.get("et",  [None, None]) or [None, None]
        pn  = match.get("pen", [None, None]) or [None, None]
        h, a = match.get("team1"), match.get("team2")
        if ft[0] is None or ft[1] is None:
            return None
        if ft[0] > ft[1]: return h
        if ft[1] > ft[0]: return a
        if et[0] is not None and et[1] is not None:
            if et[0] > et[1]: return h
            if et[1] > et[0]: return a
        if pn[0] is not None and pn[1] is not None:
            return h if pn[0] > pn[1] else a
        return None

    def _loser(match):
        w = _winner(match)
        if w is None:
            return None
        return match.get("team2") if w == match.get("team1") else match.get("team1")

    final_match  = knockout_by_round["final"][0]  if knockout_by_round["final"]  else {}
    bronze_match = knockout_by_round["bronze"][0] if knockout_by_round["bronze"] else {}

    champion   = _winner(final_match)
    runner_up  = _loser(final_match)
    third_place = _winner(bronze_match)

    return {
        "group_stage":    group_stage,
        "knockout":       knockout_flat,
        "world_champion": champion,
        "runner_up":      runner_up,
        "third_place":    third_place,
        "warnings":       warnings,
    }


# ---------------------------------------------------------------------------
# Main scan
# ---------------------------------------------------------------------------
def scan(xlsx_dir: str = None, output_path: str = None,
         fixture_json: str = None) -> None:
    root = project_root()

    xlsx_dir     = Path(xlsx_dir)     if xlsx_dir     else root / "xlsx"
    output_path  = Path(output_path)  if output_path  else root / "gData" / "predictions.json"
    fixture_json = Path(fixture_json) if fixture_json else root / "gData" / "wc2026.json"

    if not fixture_json.exists():
        raise FileNotFoundError(
            f"Fixture JSON not found: {fixture_json}\n"
            "Run extract_wc2026.py first."
        )

    with open(fixture_json, encoding="utf-8") as fh:
        wc_data = json.load(fh)

    match_index = {m["match_id"]: {"home": m["home"], "away": m["away"]}
                   for m in wc_data.get("group_stage", [])}

    xlsx_files = sorted(xlsx_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No xlsx files found in {xlsx_dir}")
        return

    players = {}
    for path in xlsx_files:
        player_name = path.stem
        print(f"  Scanning {player_name}...", end=" ")
        try:
            data = read_player_file(path, match_index)
            n_warn = len(data["warnings"])
            if n_warn:
                print(f"{n_warn} warning(s)")
                for w in data["warnings"]:
                    print(f"    ⚠  {w}")
            else:
                print("OK")
            players[player_name] = data
        except Exception as exc:
            print(f"ERROR: {exc}")

    output = {
        "meta": {
            "fixture_source": str(fixture_json),
            "n_players":      len(players),
            "players":        sorted(players.keys()),
        },
        "players": players,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(output, fh, ensure_ascii=False, indent=2)

    print(f"\nWrote {len(players)} player(s) → {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    xlsx_arg    = sys.argv[1] if len(sys.argv) > 1 else None
    output_arg  = sys.argv[2] if len(sys.argv) > 2 else None
    fixture_arg = sys.argv[3] if len(sys.argv) > 3 else None
    scan(xlsx_arg, output_arg, fixture_arg)
