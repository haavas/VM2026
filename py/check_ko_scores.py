"""
check_ko_scores.py  --  Verify knockout score encoding in participant xlsx files.

Expected (cumulative):   FT=1-1, ET=2-1, PEN=None
Wrong (per-period):      FT=1-1, ET=1-0, PEN=None

Rules:
  - ET goals must be >= FT goals for both teams (cumulative, not additive)
  - PEN goals must be >= ET goals for both teams
  - ET should only be present if FT is a draw
  - PEN should only be present if ET is a draw

Usage:
    python check_ko_scores.py [xlsx_dir]
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import column_index_from_string

ROUNDS = [
    ("r32",    "BL", "BM", "BN", "BO", list(range(10, 72, 4))),
    ("r16",    "BS", "BT", "BU", "BV", list(range(12, 48, 8))),
    ("qf",     "BZ", "CA", "CB", "CC", list(range(16, 56, 16))),
    ("sf",     "CG", "CH", "CI", "CJ", [23, 39]),
    ("final",  "CN", "CO", "CP", "CQ", [37]),
    ("bronze", "CN", "CO", "CP", "CQ", [48]),
]

def _v(ws, row, col_idx):
    v = ws.cell(row=row, column=col_idx).value
    return int(v) if isinstance(v, (int, float)) and v == int(v) else v

def check_file(path):
    issues = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2026 World Cup"]

    for (rnd, tc, ftc, etc, penc, first_rows) in ROUNDS:
        ci = {k: column_index_from_string(k) for k in [tc, ftc, etc, penc]}
        for r in first_rows:
            t1  = ws.cell(row=r,   column=ci[tc]).value
            t2  = ws.cell(row=r+1, column=ci[tc]).value
            fh  = _v(ws, r,   ci[ftc]);  fa  = _v(ws, r+1, ci[ftc])
            eh  = _v(ws, r,   ci[etc]);  ea  = _v(ws, r+1, ci[etc])
            ph  = _v(ws, r,   ci[penc]); pa  = _v(ws, r+1, ci[penc])

            # Skip empty matches
            if fh is None and fa is None:
                continue

            label = f"{rnd} row {r} ({t1} vs {t2})"

            # ET present but FT not a draw
            if (eh is not None or ea is not None) and fh != fa:
                issues.append(f"{label}: ET present but FT={fh}-{fa} is not a draw")

            # PEN present but ET not a draw
            if (ph is not None or pa is not None):
                if eh is None or ea is None:
                    issues.append(f"{label}: PEN present but no ET score")
                elif eh != ea:
                    issues.append(f"{label}: PEN present but ET={eh}-{ea} is not a draw")

            # ET scores not cumulative (must be >= FT)
            if eh is not None and fh is not None:
                if eh < fh or ea < fa:
                    issues.append(
                        f"{label}: ET={eh}-{ea} < FT={fh}-{fa} — likely per-period, not cumulative"
                    )

            # PEN scores not cumulative (must be >= ET)
            if ph is not None and eh is not None:
                if ph < eh or pa < ea:
                    issues.append(
                        f"{label}: PEN={ph}-{pa} < ET={eh}-{ea} — likely per-period, not cumulative"
                    )

    return issues


def main():
    xlsx_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("xlsx")
    files = sorted(xlsx_dir.glob("*.xlsx"))
    if not files:
        print(f"No xlsx files found in {xlsx_dir}")
        sys.exit(1)

    any_issues = False
    for f in files:
        try:
            issues = check_file(f)
        except Exception as e:
            print(f"[ERROR] {f.name}: {e}")
            continue
        if issues:
            any_issues = True
            print(f"\n{f.name}:")
            for iss in issues:
                print(f"  - {iss}")

    if not any_issues:
        print("All files OK — no inconsistencies found.")


if __name__ == "__main__":
    main()
